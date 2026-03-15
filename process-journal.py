#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
import json
import sys
from datetime import datetime

def read_excel_journal(file_path):
    """قراءة ملف Excel وتحويله إلى JSON"""
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        trades = []
        headers = []
        
        # قراءة رؤوس الأعمدة
        for cell in ws[1]:
            headers.append(cell.value)
        
        print(f"📖 الأعمدة: {headers}\n")
        
        # قراءة الصفقات
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            trade_data = {}
            
            for col_idx, cell in enumerate(row):
                header = headers[col_idx] if col_idx < len(headers) else None
                value = cell.value
                
                if header and value is not None:
                    trade_data[header] = value
            
            # تخطي الصفوف الفارغة
            if not trade_data or all(v is None for v in trade_data.values()):
                continue
            
            # معالجة البيانات
            processed = process_trade(trade_data, row_idx)
            if processed:
                trades.append(processed)
        
        return trades
    
    except Exception as e:
        print(f"❌ خطأ في قراءة الملف: {e}")
        return []

def process_trade(raw_data, row_num):
    """معالجة بيانات الصفقة الواحدة"""
    try:
        trade = {}
        
        # التاريخ
        date_val = raw_data.get('التاريخ') or raw_data.get('Date') or raw_data.get('تاريخ')
        if date_val:
            if isinstance(date_val, datetime):
                trade['date'] = date_val.isoformat()
            else:
                trade['date'] = str(date_val)
        else:
            trade['date'] = datetime.now().isoformat()
        
        # الاستراتيجية
        strategy = raw_data.get('الاستراتيجية') or raw_data.get('Strategy') or raw_data.get('استراتيجية')
        if strategy:
            strategy_map = {
                'smc': 'smc-scalping',
                'scalping': 'smc-scalping',
                'day': 'day-trading',
                'swing': 'swing-trading',
                'rejection': 'rejection',
                'fvg': 'fvg-recovery',
                'ترفع': 'rejection',
                'فجوة': 'fvg-recovery'
            }
            strategy_lower = str(strategy).lower()
            trade['strategy'] = next((v for k, v in strategy_map.items() if k in strategy_lower), 'unknown')
        
        # الرمز
        symbol = raw_data.get('الرمز') or raw_data.get('Symbol') or raw_data.get('رمز')
        trade['symbol'] = str(symbol).upper() if symbol else 'UNKNOWN'
        
        # النوع
        trade_type = raw_data.get('النوع') or raw_data.get('Type') or raw_data.get('نوع')
        trade['type'] = 'BUY' if str(trade_type).upper() in ['B', 'BUY', 'شراء', 'ش'] else 'SELL'
        
        # الأسعار
        trade['entryPrice'] = float(raw_data.get('الدخول') or raw_data.get('Entry') or 0)
        trade['stopLoss'] = float(raw_data.get('SL') or raw_data.get('Stop Loss') or 0)
        trade['takeProfit'] = float(raw_data.get('TP') or raw_data.get('Take Profit') or 0)
        trade['exitPrice'] = float(raw_data.get('الخروج') or raw_data.get('Exit') or 0)
        
        # النتيجة
        result = raw_data.get('النتيجة') or raw_data.get('Result') or raw_data.get('نتيجة')
        if result:
            result_str = str(result).lower()
            if 'win' in result_str or 'w' in result_str or '+' in result_str:
                trade['result'] = 'win'
            elif 'loss' in result_str or 'l' in result_str or '-' in result_str:
                trade['result'] = 'loss'
            elif 'break' in result_str or 'even' in result_str or '0' in result_str:
                trade['result'] = 'breakeven'
            else:
                trade['result'] = 'unknown'
        else:
            trade['result'] = 'unknown'
        
        # حساب PnL إذا لم يكن موجود
        pnl_val = raw_data.get('PnL') or raw_data.get('الربح/الخسارة')
        if pnl_val:
            trade['pnl'] = float(pnl_val)
        else:
            if trade['entryPrice'] and trade['exitPrice']:
                trade['pnl'] = trade['exitPrice'] - trade['entryPrice']
            else:
                trade['pnl'] = 0
        
        # المدة
        duration = raw_data.get('المدة') or raw_data.get('Duration') or 0
        trade['duration'] = int(duration) if duration else 0
        
        # الأخطاء النفسية
        mistakes = raw_data.get('الأخطاء') or raw_data.get('Mistakes') or ''
        if mistakes:
            mistake_list = []
            mistakes_str = str(mistakes).lower()
            if 'fomo' in mistakes_str:
                mistake_list.append('fomo')
            if 'revenge' in mistakes_str:
                mistake_list.append('revenge')
            if 'early' in mistakes_str:
                mistake_list.append('early-exit')
            if 'oversize' in mistakes_str or 'size' in mistakes_str:
                mistake_list.append('oversizing')
            if 'break' in mistakes_str and 'sl' in mistakes_str:
                mistake_list.append('breaking-sl')
            trade['mistakes'] = mistake_list
        else:
            trade['mistakes'] = []
        
        # الملاحظات
        trade['notes'] = str(raw_data.get('ملاحظات') or raw_data.get('Notes') or '')
        
        # درجة نفسية
        trade['psychScore'] = 100 - (len(trade['mistakes']) * 15)
        
        return trade
    
    except Exception as e:
        print(f"⚠️ تحذير في الصف {row_num}: {e}")
        return None

def save_trades_json(trades, output_file):
    """حفظ الصفقات في JSON"""
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(trades, f, ensure_ascii=False, indent=2)
    print(f"✅ تم حفظ {len(trades)} صفقة في {output_file}")

def print_trades_summary(trades):
    """طباعة ملخص الصفقات"""
    if not trades:
        print("❌ لا توجد صفقات")
        return
    
    print(f"\n📊 ملخص البيانات:")
    print(f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
    print(f"إجمالي الصفقات: {len(trades)}")
    
    # إحصائيات سريعة
    wins = len([t for t in trades if t['result'] == 'win'])
    losses = len([t for t in trades if t['result'] == 'loss'])
    
    print(f"الفوزة: {wins}")
    print(f"الخسائر: {losses}")
    
    if wins + losses > 0:
        wr = (wins / (wins + losses) * 100)
        print(f"Win Rate: {wr:.1f}%")
    
    # إحصائيات الاستراتيجيات
    strategies = {}
    for trade in trades:
        strat = trade['strategy']
        if strat not in strategies:
            strategies[strat] = {'count': 0, 'wins': 0}
        strategies[strat]['count'] += 1
        if trade['result'] == 'win':
            strategies[strat]['wins'] += 1
    
    print(f"\n📈 الاستراتيجيات:")
    for strat, data in strategies.items():
        wr = (data['wins'] / data['count'] * 100) if data['count'] > 0 else 0
        print(f"  • {strat}: {data['count']} صفقة، {wr:.0f}% WR")
    
    # عينة من الصفقات
    print(f"\n📋 عينة من الصفقات:")
    for i, trade in enumerate(trades[:3], 1):
        print(f"  {i}. {trade['date']} | {trade['strategy']} | {trade['symbol']} | {trade['result']}")

if __name__ == "__main__":
    # قراءة الملف
    file_path = sys.argv[1] if len(sys.argv) > 1 else "/home/ubuntu/.openclaw/media/inbound/8cc0ff29-72ce-42b8-9363-985c8cef4ce8.xlsx"
    
    print(f"📖 قراءة الملف: {file_path}\n")
    
    trades = read_excel_journal(file_path)
    
    if trades:
        output_file = '/home/ubuntu/.openclaw/workspace/trades-journal.json'
        save_trades_json(trades, output_file)
        print_trades_summary(trades)
    else:
        print("❌ فشلت قراءة الصفقات")
